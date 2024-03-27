import io
from typing import Dict, List
from openpyxl import Workbook


def create_excel(headers: Dict[str, str], data: List[Dict[str, str]]) -> io.BytesIO:
    """
    :param headers:
    :param data:
    :return: io.BytesIO
    """
    output: io.BytesIO = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active

    for column, header in enumerate(headers.values()):
        sheet.cell(row=1, column=column + 1).value = header

    row_num: int = 2
    for data_dict in data:
        for column, key in enumerate(headers.keys()):
            sheet.cell(row=row_num, column=column + 1).value = data_dict.get(key, '')
        row_num += 1

    workbook.save(output)
    output.seek(0)  # 回到流的开始位置

    return output
#
#
# if __name__ == '__main__':
#     headers = {
#         'name': '姓名',
#         'age': '年龄',
#         'gender': '性别',
#         'address': '地址',
#     }
#     data = [
#         {'name': '张三', 'age': 18, 'gender': '男', 'address': '北京市'},
#         {'name': '李四', 'age': 20, 'gender': '女', 'address': '上海市'},
#         {'name': '王五', 'age': 22, 'gender': '男', 'address': '广州市'},
#         ]
#     io_data = create_excel(headers, data)
#     with open('output.xlsx', 'wb') as f:
#         f.write(io_data.getvalue())
